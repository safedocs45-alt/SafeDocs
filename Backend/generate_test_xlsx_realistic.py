# generate_test_xlsx_realistic.py
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

ws['A1'] = "SafeDocs Realistic Demo - XLSX (INERT STUBS)"
ws['A2'] = "Marker: X-SAFEDOCS-MARKER: MAL_TEST_EXCEL_MACRO_STUB"

# Insert a cell with a pseudo-macro snippet as plain text
ws['A4'] = "' PSEUDO-MACRO: inert stub for demo purposes"
ws['A5'] = "Sub Workbook_Open(): ' inert stub - DO NOT RUN"

# Hidden sheet with a placeholder name that detectors often look for
ws2 = wb.create_sheet("vbaProject_placeholder")
ws2.sheet_state = 'hidden'
ws2['A1'] = 'OLE_PAYLOAD_STUB: This is a harmless placeholder to simulate an embedded object.'

wb.properties.title = "SafeDocs Demo XLSX Realistic"
wb.save('safedocs_realistic_xlsx.xlsx')
print("Created safedocs_realistic_xlsx.xlsx")
